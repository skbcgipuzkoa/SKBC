import json
import re
from pathlib import Path

import openpyxl

ROOT = Path.cwd()
WORKBOOK_PATH = Path("C:/Users/alvar/Desktop/shorinji_kempo_club.xlsx")
SYLLABUS_PATHS = [
    Path("C:/Users/alvar/Desktop/CLAUDE/superpowers-main/skbc-gipuzkoa/exam-system/kyu-syllabus-extracted.txt"),
    Path("C:/Users/alvar/Desktop/CLAUDE/superpowers-main/skbc-gipuzkoa/exam-system/dan-syllabus-extracted.txt"),
]
STANCE_WORDS = ["Tai", "Hiraki", "Both", "Any", "All"]
SECTION_STOP = re.compile(
    r"\s+(--\s+\d+\s+of\s+24\s+--|Goho\s+Juho|Tai gamae|Hiraki gamae|Umpo ho|Ukemi|"
    r"Kata tan-en|Kata sotai|Kumi embu|Randori|Keimyaku|Gakka|Dokun|Vocab list|Notes)\b",
    re.I,
)
LAYOUT_STOP = re.compile(r"\s+((Goho|Juho)\s+\d|White|Yellow|Orange|Green|Blue|Brown|Black|Shodan|Nidan|Sandan|Yondan|Godan)\b", re.I)


def main():
    names = load_technique_names()
    syllabus = " ".join(path.read_text(encoding="utf-8", errors="replace") for path in SYLLABUS_PATHS)
    syllabus = re.sub(r"\s+", " ", syllabus.replace("\ufffd", "").replace("“", '"').replace("”", '"')).strip()
    entry_regexes = build_entry_regexes(names)
    summaries = {}
    for name in names:
        entry = find_entry(name, syllabus, entry_regexes)
        if entry:
            summaries[normalize_key(name)] = to_spanish_summary(entry)

    output = ROOT / "src/lib/kamoku-summary-fallbacks.ts"
    output.write_text(render_ts(summaries), encoding="utf-8")
    print(json.dumps({"techniques": len(names), "summaries": len(summaries), "output": str(output)}, indent=2))


def load_technique_names():
    workbook = openpyxl.load_workbook(WORKBOOK_PATH, read_only=True, data_only=True)
    sheet = workbook["TECNICAS_ADULTOS"]
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    name_index = headers.index("NOMBRE_TECNICA")
    names = {str(row[name_index]).strip() for row in sheet.iter_rows(min_row=2, values_only=True) if row[name_index]}
    return sorted(names, key=len, reverse=True)


def build_entry_regexes(names):
    extra = [
        "Gyaku ten ichi (ura & omote)",
        "Kaishin zuki (ura & omote)",
        "Uchi age zuki (ura & omote)",
        "Uchi age geri (ura & omote)",
        "Soto uke geri (ura & omote)",
        "Uchi oshi uke geri (ura & omote)",
        "Johaku nuki (katate & ryote)",
        "Johaku dori (katate & ryote)",
        "Sode nuki (katate & ryote)",
        "Uchi nuki (katate & ryote)",
        "Juji nuki (katate & ryote)",
        "Juji gote (katate & ryote)",
        "Kiri gote (katate & morote)",
        "Kiri kaeshi tembin (katate & morote)",
        "Okuri tembin dori (two types)",
        "Okuri gassho (two types)",
        "Fukko daoshi (3 types)",
        "Hasami daoshi (2 types)",
        "Tora daoshi (2 types)",
        "Ryu nage, ryu gatame",
    ]
    patterns = []
    for name in sorted(set(names + extra), key=len, reverse=True):
        patterns.append(re.compile(rf"{technique_pattern(name)}\s+({'|'.join(STANCE_WORDS)})\b", re.I))
    return patterns


def find_entry(name, syllabus, entry_regexes):
    for candidate in [name] + variant_candidates(name):
        match = re.search(rf"{technique_pattern(candidate)}\s+({'|'.join(STANCE_WORDS)})\b", syllabus, re.I)
        if not match:
            continue
        start = match.start()
        end = len(syllabus)
        for regex in entry_regexes:
            next_match = regex.search(syllabus, match.end())
            if next_match and start < next_match.start() < end:
                end = next_match.start()
        section_stop = SECTION_STOP.search(syllabus, match.end())
        if section_stop and start < section_stop.start() < end:
            end = section_stop.start()
        layout_stop = LAYOUT_STOP.search(syllabus, match.end())
        if layout_stop and start < layout_stop.start() < end:
            end = layout_stop.start()
        raw = re.sub(r"\s+", " ", syllabus[start:end]).strip()
        if len(raw) > 12:
            return raw
    return ""


def variant_candidates(name):
    candidates = []
    base = re.sub(r"^(Katate|Morote|Ryote|Ura|Omote)\s+", "", name, flags=re.I)
    base = re.sub(r"\s+\((ura|omote|katate|ryote|two types|3 types)\)", "", base, flags=re.I).strip()
    if base and base != name:
        candidates.append(base)
    grouped = [
        ("gyaku ten ichi", "Gyaku ten ichi (ura & omote)"),
        ("kaishin zuki", "Kaishin zuki (ura & omote)"),
        ("uchi age zuki", "Uchi age zuki (ura & omote)"),
        ("uchi age geri", "Uchi age geri (ura & omote)"),
        ("soto uke geri", "Soto uke geri (ura & omote)"),
        ("uchi oshi uke geri", "Uchi oshi uke geri (ura & omote)"),
        ("johaku nuki", "Johaku nuki (katate & ryote)"),
        ("johaku dori", "Johaku dori (katate & ryote)"),
        ("sode nuki", "Sode nuki (katate & ryote)"),
        ("uchi nuki", "Uchi nuki (katate & ryote)"),
        ("juji nuki", "Juji nuki (katate & ryote)"),
        ("juji gote", "Juji gote (katate & ryote)"),
        ("kiri gote", "Kiri gote (katate & morote)"),
        ("kiri kaeshi tembin", "Kiri kaeshi tembin (katate & morote)"),
        ("okuri tembin dori", "Okuri tembin dori (two types)"),
        ("okuri gassho", "Okuri gassho (two types)"),
        ("fukko daoshi", "Fukko daoshi (3 types)"),
        ("hasami daoshi", "Hasami daoshi (2 types)"),
        ("tora daoshi", "Tora daoshi (2 types)"),
        ("ryu nage", "Ryu nage, ryu gatame"),
    ]
    lowered = name.lower()
    for needle, candidate in grouped:
        if needle in lowered:
            candidates.append(candidate)
    return candidates


def to_spanish_summary(raw):
    parsed = re.match(rf"^(.*?)\s+({'|'.join(STANCE_WORDS)})\s+(.*)$", raw, re.I)
    if not parsed:
        return translate(raw)
    stance = parsed.group(2)
    detail = parsed.group(3)
    detail = re.sub(r"\bA:\s*", "Atacante: ", detail)
    detail = re.sub(r"\bD:\s*", "Defensor: ", detail)
    return f"Posicion: {translate_stance(stance)}. {translate(detail)}"


def translate(value):
    replacements = [
        (r"\bAttacker:\s*", "Atacante: "),
        (r"\bDefender:\s*", "Defensor: "),
        (r"\bAs for\b", "Como en"),
        (r"\bAlso from\b", "Tambien desde"),
        (r"\bAdaptation of\b", "Adaptacion de"),
        (r"\bFrom\b", "Desde"),
        (r"\battack\b", "atacar"),
        (r"\battempt\b", "intento de"),
        (r"\bgrab both wrists\b", "agarra ambas muñecas"),
        (r"\bgrab both lapels\b", "agarra ambas solapas"),
        (r"\bgrab lapel\b", "agarra la solapa"),
        (r"\bgrab sleeve\b", "agarra la manga"),
        (r"\bcross grab lower sleeve\b", "agarre cruzado de manga baja"),
        (r"\bcross grab sleeve \(upper\)", "agarre cruzado de manga alta"),
        (r"\bcross grab inside wrist\b", "agarre cruzado por dentro de la muñeca"),
        (r"\bsame side grab outside wrist\b", "agarre del mismo lado por fuera de la muñeca"),
        (r"\bgrab upper sleeve\(s\) and pull\b", "agarra una o ambas mangas altas y tira"),
        (r"\bgrab bicep\(s\) and push\b", "agarra uno o ambos brazos por el biceps y empuja"),
        (r"\bgrab wrist first\b", "agarra primero la muñeca"),
        (r"\bgrab wrist\b", "agarra la muñeca"),
        (r"\btwo attackers, one person grab each arm & pull\b", "dos atacantes, cada uno agarra un brazo y tira"),
        (r"\bjudo grip\b", "agarre tipo judo"),
        (r"\bsumo grip\b", "agarre tipo sumo"),
        (r"\bhandshake\b", "saludo de manos"),
        (r"\bpush away\b", "empuja alejando"),
        (r"\bpush\b", "empuja"),
        (r"\bpull back\b", "retrocede tirando"),
        (r"\bpull\b", "tira"),
        (r"\brelease grip\b", "suelta el agarre"),
        (r"\btwist hand outwards to resist\b", "gira la mano hacia fuera para resistir"),
        (r"\bstrike wrist\b", "golpea la muñeca"),
        (r"\bstrike\b", "golpea"),
        (r"\boffer raised hand\b", "ofrece la mano levantada"),
        (r"\binterlace fingers\b", "entrelaza los dedos"),
        (r"\btwist in\b", "gira hacia dentro"),
        (r"\bupper\b", "alta"),
        (r"\blower\b", "baja"),
        (r"\bthen\b", "despues"),
        (r"\bfirst\b", "primero"),
        (r"\bstraight\b", "directo"),
        (r"\bhigh lapel & sleeve\b", "solapa alta y manga"),
        (r"\bprotectors\b", "protecciones"),
        (r"\bpin\b", "control"),
        (r"\bthrow\b", "proyeccion"),
        (r"\bstab down from above\b", "cuchillada descendente desde arriba"),
        (r"\bstab chudan\b", "cuchillada chudan"),
        (r"\bwalk past & grab wrist\b", "pasa caminando y agarra la muñeca"),
        (r"\bturn\b", "girar"),
        (r"\btake along\b", "conduccion"),
        (r"\bstanding arm bar\b", "luxacion de brazo de pie"),
        (r"\bthrow away\b", "proyecta alejando"),
        (r"\buse neck\b", "usa el cuello"),
        (r"\buse one hand\b", "usa una mano"),
        (r"\bfront hand\b", "mano adelantada"),
        (r"\bboth wrists\b", "ambas muñecas"),
        (r"\bboth lapels\b", "ambas solapas"),
        (r"\btwo hands\b", "dos manos"),
        (r"\btwo types\b", "dos formas"),
        (r"\bthree types\b", "tres formas"),
        (r"\boutside\b", "por fuera"),
        (r"\binside\b", "por dentro"),
        (r"\bsame side\b", "mismo lado"),
        (r"\bcross\b", "cruzado"),
        (r"\belbows out\b", "codos hacia fuera"),
        (r"\bpalms or thumbs\b", "palmas o pulgares"),
        (r"\bopen\b", "abierto"),
        (r"\bwith\b", "con"),
        (r"\bto\b", "a"),
        (r"\bor\b", "o"),
        (r"\band\b", "y"),
        (r"\bplus\b", "mas"),
        (r"\bagainst\b", "contra"),
        (r"\bdefence\b", "defensa"),
        (r"\bwhen\b", "cuando"),
        (r"\bfails\b", "falla"),
        (r"\bclose in\b", "entra cerrando distancia"),
        (r"\bextended\b", "extendido"),
        (r"\bfrom\b", "desde"),
        (r"\bany punch\b", "cualquier golpe de puño"),
        (r"\bsequence\b", "secuencia"),
    ]
    text = value
    for pattern, replacement in replacements:
        text = re.sub(pattern, replacement, text, flags=re.I)
    text = re.sub(r"\s+(Goho|Juho)\s+(--\s+\d+\s+of\s+24\s+--.*|\d+dan.*|st.*|nd.*|rd.*|th.*)$", "", text, flags=re.I)
    text = re.sub(r"\s+(Goho|Juho)\s*$", "", text, flags=re.I)
    return re.sub(r"\s+([.,;:])", r"\1", text).strip()


def translate_stance(stance):
    return {
        "tai": "tai gamae",
        "hiraki": "hiraki gamae",
        "both": "tai/hiraki segun ura u omote",
        "any": "cualquier kamae indicado",
        "all": "todas las formas indicadas",
    }.get(stance.lower(), stance)


def render_ts(summaries):
    body = json.dumps(summaries, ensure_ascii=False, indent=2)
    return (
        "// Generated from the local Kyu/Dan syllabus extracts. Run scripts/generate-kamoku-summary-fallbacks.py to refresh.\n"
        f"export const kamokuSummaryFallbacks: Record<string, string> = {body};\n\n"
        "export function getKamokuSummaryFallback(name: string | null | undefined) {\n"
        "  const key = normalizeKamokuKey(name);\n"
        "  if (!key) return \"\";\n"
        "  return kamokuSummaryFallbacks[key] ?? kamokuSummaryFallbacks[stripKamokuVariantKey(key)] ?? \"\";\n"
        "}\n\n"
        "export function normalizeKamokuKey(value: string | null | undefined) {\n"
        "  return String(value ?? \"\").trim().toLowerCase().normalize(\"NFD\").replace(/[\\u0300-\\u036f]/g, \"\").replace(/\\s+/g, \" \");\n"
        "}\n\n"
        "function stripKamokuVariantKey(key: string) {\n"
        "  return key\n"
        "    .replace(/^(katate|morote|ryote|ura|omote) /, \"\")\n"
        "    .replace(/ \\((ura|omote|katate|ryote|two types|3 types)\\)/g, \"\")\n"
        "    .replace(/\\s+/g, \" \")\n"
        "    .trim();\n"
        "}\n"
    )


def technique_pattern(name):
    return re.escape(name).replace(r"\ ", r"\s+")


def normalize_key(value):
    text = str(value or "").strip().lower()
    return re.sub(r"\s+", " ", text)


if __name__ == "__main__":
    main()
