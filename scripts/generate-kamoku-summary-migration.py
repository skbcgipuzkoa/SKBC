import importlib.util
import re
from pathlib import Path

import openpyxl

ROOT = Path.cwd()
GENERATOR_PATH = ROOT / "scripts" / "generate-kamoku-summary-fallbacks.py"
WORKBOOK_PATH = Path("C:/Users/alvar/Desktop/shorinji_kempo_club.xlsx")
OUTPUT_PATH = ROOT / "supabase" / "migrations" / "20260827090000_refresh_kamoku_summaries.sql"


def main():
    generator = load_generator()
    names = generator.load_technique_names()
    syllabus = " ".join(
        path.read_text(encoding="utf-8", errors="replace")
        for path in generator.SYLLABUS_PATHS
    )
    syllabus = re.sub(r"\s+", " ", syllabus.replace("\ufffd", "")).strip()
    entry_regexes = generator.build_entry_regexes(names)

    rows = []
    workbook = openpyxl.load_workbook(WORKBOOK_PATH, read_only=True, data_only=True)
    sheet = workbook["TECNICAS_ADULTOS"]
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    legacy_index = headers.index("ID_TECNICA")
    name_index = headers.index("NOMBRE_TECNICA")

    for row in sheet.iter_rows(min_row=2, values_only=True):
      legacy_id = str(row[legacy_index] or "").strip()
      name = str(row[name_index] or "").strip()
      if not legacy_id or not name:
          continue
      entry = generator.find_entry(name, syllabus, entry_regexes)
      summary = generator.to_spanish_summary(entry) if entry else ""
      if summary:
          rows.append((legacy_id, summary))

    OUTPUT_PATH.write_text(render_sql(rows), encoding="utf-8")
    print({"rows": len(rows), "output": str(OUTPUT_PATH)})


def load_generator():
    spec = importlib.util.spec_from_file_location("kamoku_generator", GENERATOR_PATH)
    module = importlib.util.module_from_spec(spec)
    assert spec.loader
    spec.loader.exec_module(module)
    return module


def render_sql(rows):
    values = ",\n  ".join(
        f"('{escape_sql(legacy_id)}', '{escape_sql(summary)}')" for legacy_id, summary in rows
    )
    return f"""-- Refresh adult technique summaries from the BSKF kyu/dan syllabus extracts.
-- The system keeps split SKBC technique rows by ID_TECNICA and applies the matching syllabus summary.

with summaries(legacy_id, summary_es) as (
  values
  {values}
)
update public.techniques t
set summary_es = summaries.summary_es,
    updated_at = now()
from summaries
where t.legacy_id = summaries.legacy_id;

update public.technical_plans tp
set summary_es = t.summary_es,
    updated_at = now()
from public.techniques t
where tp.technique_id = t.id
  and t.summary_es is not null;
"""


def escape_sql(value):
    return str(value).replace("'", "''")


if __name__ == "__main__":
    main()
